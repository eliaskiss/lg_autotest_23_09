import sys

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles.fonts import Font
from openpyxl.styles.borders import Border, Side

# Path의 맨마지막에 추가
# sys.path.append('D:\source\lg_autotest_23_08\Elias\3일차\lec_pymysql')
sys.path.append('..\lec_pymysql')

# 우선순위 0번째 순위로 Path 추가
# sys.path.insert(0, '..\lec_pymysql')

# 상위폴더 Path 추가
# sys.path.append('..')

from lec_pymysql import Database

# Create Table SQL
# table_name 지정필요!!!!
# CREATE TABLE `bicycle` (
#   `id` INT(11) NOT NULL AUTO_INCREMENT,
#   `reg_datetime` DATETIME DEFAULT CURRENT_TIMESTAMP(),
#   `station_number` INT(11) DEFAULT NULL,
#   `station_name` VARCHAR(128) DEFAULT NULL,
#   `region` VARCHAR(128) DEFAULT NULL,
#   `address` VARCHAR(1024) DEFAULT NULL,
#   `latitude` FLOAT DEFAULT NULL,
#   `longitude` FLOAT DEFAULT NULL,
#   `install_date` DATETIME DEFAULT NULL,
#   `lcd_count` INT(11) DEFAULT NULL,
#   `qr_count` INT(11) DEFAULT NULL,
#   `proc_type` VARCHAR(128) DEFAULT NULL,
#   KEY `id` (`id`)
# ) ENGINE=INNODB DEFAULT CHARSET=utf8 COLLATE=utf8_general_ci;

# Task1
# public_bicycle.xlsx 파일을 읽어서, DB의 public_bycycle의 테이블에 넣기

DB_URL = '45.115.155.124'
DB_USER = 'dbadmin'
DB_PW = 'dbadmin'
DB_NAME = 'bicycle'

def put_data_to_db(excel_file_name):
    # Load wb from excel file
    wb = load_workbook(excel_file_name, data_only=True, read_only=True)

    # Select work sheet
    ws = wb.active
    # ws = wb['대여소현황']

    # DB 객체 생성
    db = Database(DB_URL, DB_USER, DB_PW, DB_NAME)

    # DB 연결
    db.connect_db()

    columns = ['station_number', 'station_name', 'region', 'address', 'latitude', 'longitude', 'install_date', 'lcd_count',
               'qr_count', 'proc_type']

    # Read data from work sheet
    for row in ws.iter_rows(min_row=6):
        # station_number = row[0].value
        # station_name = row[1].value
        # region = row[2].value
        # address = row[3].value
        # latitude = row[4].value
        # longitude = row[5].value
        # install_date = row[6].value
        # lcd_count = row[7].value
        # qr_count = row[8].value
        # proc_type = row[9].value
        #
        # # Way I
        # sql = 'insert into elias (station_number, station_name, region, address, latitude, longitude,' \
        #       'install_date, lcd_count, qr_count, proc_type) values(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s);'
        # values = (station_number, station_name, region, address, latitude, longitude,
        #           install_date, lcd_count, qr_count, proc_type)
        # db.execute_only(sql, values)

        # # Way II
        # sql = 'insert into elias (station_number, station_name, region, address, latitude, longitude,' \
        #       'install_date, lcd_count, qr_count, proc_type) values(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s);'
        # values = [_row.value for _row in row]
        # db.execute_only(sql, values)

        # Way III
        sql = f'insert into elias ({",".join(columns)}) values({("%s," * len(columns))[:-1]});'
        values = [_row.value for _row in row]
        db.execute_only(sql, values)



    # DB에 Commit
    db.commit_only()

    # DB 연결해제
    db.disconnect_db()

# Task2
# DB의 bicycle의 테이블에서 특정 데이터를 뽑아서, 엑셀로 저장하기
# 2020이후에 서초구에 설치된 자전거 대여소 목록데이터
# sql = 'SELECT * FROM elias WHERE DATE(install_date) >= "2020-01-01" AND region = "서초구";'

# sql = 'SELECT * FROM elias WHERE DATE(install_date) >= %s AND region = %s;'
# from_date = "2020-01-01"
# region = "서초구"
# values = (from_date, region)
# execute(sql, values)

def get_data_from_db(from_date, region, output_file_name):
    # Create Workbook
    wb = Workbook()

    # Select Worksheet
    ws = wb.active

    # Rename Worksheet
    ws.title = '공유자전거'

    # Header Setting (Merge)
    ws['A1'] = '대여소\n번호'
    ws.merge_cells('A1:A5')

    ws['B1'] = '보관소(대여소)명'
    ws.merge_cells('B1:B5')

    ws['C1'] = '소재자(위치)'
    ws.merge_cells('C1:F2')

    ws['C3'] = '자치구'
    ws.merge_cells('C3:C5')

    ws['D3'] = '상세주소'
    ws.merge_cells('D3:D5')

    ws['E3'] = '위도'
    ws.merge_cells('E3:E5')

    ws['F3'] = '경도'
    ws.merge_cells('F3:F5')

    ws['G1'] = '설치시기'
    ws.merge_cells('G1:G5')

    ws['H1'] = '설치형태'
    ws.merge_cells('H1:I1')

    ws['H2'] = 'LCD'
    ws.merge_cells('H2:H3')

    ws['H4'] = '거치대수'
    ws.merge_cells('H4:H5')

    ws['I2'] = 'QR'
    ws.merge_cells('I2:I3')

    ws['I4'] = '거치대수'
    ws.merge_cells('I4:I5')

    ws['J1'] = '운영방식'
    ws.merge_cells('J1:J5')

    # Get data from DB with sql = 'select * from elias where date(install_date) >= "2020-01-01" and region = "서초구";'
    db = Database(DB_URL, DB_USER, DB_PW, DB_NAME)
    db.connect_db()

    sql = 'SELECT * FROM elias WHERE DATE(install_date) >= %s AND region = %s;'
    values = (from_date, region)
    data_list = db.execute_and_return(sql, values)

    # todo: DB로 부터 가져온 데이터를 Excel에 저장
    for data in data_list:
        # Way I
        # ws.append([data['station_number'], data['station_name'], data['install_date'].strftime()...])

        # Way II
        ws.append([_data for _data in list(data.values())[2:]])

    # Add Stlying
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    bold_white_font = Font(bold=True,
                           size=12,
                           italic=None,
                           underline=None,
                           strike=None,
                           color='FFFFFF')

    background_fill = PatternFill(start_color='525E75',
                                  end_color='525E75',
                                  fill_type='solid')

    # Header Style 적용 : 백그라운드 색 + 화이트 Bold 폰트
    for row in ws.iter_rows(max_row=5):
        for cell in row:
            cell.fill = background_fill
            cell.font = bold_white_font

    # 테두리 적용 : thin border
    for col in ws.columns:
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            cell.border = thin_border

    wb.save(output_file_name)


if __name__ == '__main__':
    # put_data_to_db('public_bicycle.xlsx')
    get_data_from_db('2020-01-01', '서초구', 'new_excel.xlsx')

