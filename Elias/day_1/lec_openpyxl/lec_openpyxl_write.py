from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from datetime import datetime, timedelta

############################################################
# Create new excel workbook
############################################################
wb = Workbook()

file_name = './new_excel.xlsx'

############################################################
# Get Active Worksheet
############################################################
ws = wb.active

############################################################
# Set Sheet Name
############################################################
ws.title = 'Basic'

############################################################
# Set Cell Value
############################################################
ws['A1'] = 'Hello World'
ws['B1'] = 10
ws['C1'] = 20

############################################################
# Use formular
############################################################
ws['D1'] = '=SUM(B1+C1)'

############################################################
# Set Sheet Name
############################################################
comment = Comment('This is comment', 'Elias Kim', 100, 100)
ws['A1'].comment = comment

############################################################
# Insert Row
############################################################
header = ['V1', 'V2', 'V3', 'V4', 'V5']
ws.append(header)
row = [10, 20, 30, 40, 50]
for i in range(10):
    ws.append(row)

############################################################
# Insert Datetime
############################################################
ws['A14'] = datetime.now()
ws['B14'] = datetime.now() + timedelta(days=1) # 내일
ws['C14'] = datetime.now() - timedelta(days=1) # 어제
ws['D14'] = datetime.now() + timedelta(hours=1) # 한시간뒤
ws['E14'] = '2023-04-12'

############################################################
# Merge/Unmerge Target Cell
############################################################
ws['A16'] = 'Hello'
ws['B16'] = 'World'

# Way I
# ws.merge_cells(range_string='A16:B16')
# ws.unmerge_cells('A16:B16')

# Way II
ws.merge_cells(start_row=16, start_column=1, end_row=16, end_column=2)
ws.unmerge_cells(start_row=16, start_column=1, end_row=16, end_column=2)

############################################################
# Insert Image
############################################################
img = Image('buz.jpg')
ws.add_image(img, 'G1')

############################################################
# Create 2nd Sheet
############################################################
wb.create_sheet('Insert Delete Move')
wb.active = 1

############################################################
# Select Sheet
############################################################
ws = wb['Insert Delete Move']

############################################################
# Create Rows Data
############################################################
for i in range(20):
    new_row = [f'{chr(ord("A") + j)}{i+1}' for j in range(20)] # [A1, B1, ..., T1] [A2, B2, C2, ..., T2]
    ws.append(new_row)

# ############################################################
# # Insert Emtpy Row
# ############################################################
# ws.insert_rows(2)
#
# ############################################################
# # Insert Emtpy Column
# ############################################################
# ws.insert_cols(2)

############################################################
# Delete Rows
############################################################
ws.delete_rows(5, 7) # From 5, Step 7

############################################################
# Delete Columns
############################################################
ws.delete_cols(5, 7) # From 5, Step 7

############################################################
# Move Row
############################################################
ws.move_range('A1:T1', rows=21, cols=1) # 'A1:U1', Row step 21, Column step 1

############################################################
# Create 3rd Sheet
# Doc Site: https://openpyxl.readthedocs.io/en/stable/charts/introduction.html
############################################################
wb.create_sheet('Chart')
wb.active = 2
ws = wb['Chart']

from openpyxl.chart import (
    AreaChart,
    AreaChart3D,
    Reference,
    BubbleChart,
    BarChart,
    Series
)
from copy import deepcopy

ws['A1'] = 'Bar Chart'
rows = [
    ('Number', 'Batch 1', 'Batch 2'),
    (2, 10, 30),
    (3, 40, 60),
    (4, 50, 70),
    (5, 20, 10),
    (6, 10, 40),
    (7, 50, 30),
]

for row in rows:
    ws.append(row)

############################################################################
# Bar Chart
############################################################################

chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Bar Chart"
chart1.y_axis.title = 'Test number'
chart1.x_axis.title = 'Sample length (mm)'

data = Reference(ws, min_col=2, min_row=2, max_row=8, max_col=3)
cats = Reference(ws, min_col=1, min_row=3, max_row=8)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4
ws.add_chart(chart1, "E1")

############################################################################
# Horizontal Bar Chart
############################################################################
chart2 = deepcopy(chart1)
chart2.style = 11
chart2.type = "bar"
chart2.title = "Horizontal Bar Chart"

ws.add_chart(chart2, "N1")

############################################################################
# Stack Bar Chart
############################################################################
chart3 = deepcopy(chart1)
chart3.type = "col"
chart3.style = 12
chart3.grouping = "stacked"
chart3.overlap = 100
chart3.title = 'Stacked Chart'

ws.add_chart(chart3, "E15")

############################################################################
# Horizontal Stack Bar Chart
############################################################################
chart4 = deepcopy(chart1)
chart4.type = "bar"
chart4.style = 13
chart4.grouping = "percentStacked"
chart4.overlap = 100
chart4.title = 'Percent Stacked Chart'

ws.add_chart(chart4, "N15")

############################################################################
# Area Chart
############################################################################
ws['A29'] = 'Area Chart'

rows = [
    ['Number', 'Batch 1', 'Batch 2'],
    [2, 40, 30],
    [3, 40, 25],
    [4, 50, 30],
    [5, 30, 10],
    [6, 25, 5],
    [7, 50, 10],
]

for row in rows:
    ws.append(row)

chart = AreaChart()
chart.title = "Area Chart"
chart.style = 13
chart.x_axis.title = 'Test'
chart.y_axis.title = 'Percentage'

cats = Reference(ws, min_col=1, min_row=1, max_row=7)
data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=7)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

ws.add_chart(chart, "E29")

############################################################################
# Area 3D Chart
############################################################################
chart = AreaChart3D()
chart.title = "Area Chart"
chart.style = 13
chart.x_axis.title = 'Test'
chart.y_axis.title = 'Percentage'
chart.legend = None

cats = Reference(ws, min_col=1, min_row=1, max_row=7)
data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=7)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

ws.add_chart(chart, "N29")

############################################################################
# Bubble Chart
############################################################################
ws['A43'] = 'Bubble Chart'
rows = [
    ("Number of Products", "Sales in USD", "Market share"),
    (14, 12200, 15),
    (20, 60000, 33),
    (18, 24400, 10),
    (22, 32000, 42),
    (),
    (12, 8200, 18),
    (15, 50000, 30),
    (19, 22400, 15),
    (25, 25000, 50),
]

for row in rows:
    ws.append(row)

chart = BubbleChart()
chart.style = 18 # use a preset style

# add the first series of data
xvalues = Reference(ws, min_col=1, min_row=31, max_row=34)
yvalues = Reference(ws, min_col=2, min_row=31, max_row=34)
size = Reference(ws, min_col=3, min_row=31, max_row=34)
series = Series(values=yvalues, xvalues=xvalues, zvalues=size, title="2013")
chart.series.append(series)

# add the second
xvalues = Reference(ws, min_col=1, min_row=36, max_row=39)
yvalues = Reference(ws, min_col=2, min_row=36, max_row=39)
size = Reference(ws, min_col=3, min_row=36, max_row=39)
series = Series(values=yvalues, xvalues=xvalues, zvalues=size, title="2014")
chart.series.append(series)

# place the chart starting in cell E1
ws.add_chart(chart, "E43")










wb.save(file_name)



























