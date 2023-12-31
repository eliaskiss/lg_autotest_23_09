from openpyxl import load_workbook
from datetime import datetime, timedelta
from icecream import ic
import time

file_name = 'public_bicycle.xlsx'

#############################################################
# Open Excel File
#############################################################
wb = load_workbook(file_name,
                   read_only=False, # Read Only가 True이면, Lazy Loading이 발생해서 반드시 Close 해줘야만 핸들이 릴리즈됨
                   data_only=True) # Datay Only가 True이면, Cell의 함수 결과값을 가져옴

#############################################################
# Get Sheet List
#############################################################
ws_list = wb.sheetnames
# ic(ws_list)

#############################################################
# Select Worksheet
#############################################################
ws = wb.active      # 현재 활성화되어 있는 sheet(저장할때 선택된 시트)
# ws = wb['대여소현황']
# ws = wb[ws_list[0]]
# ws = wb[ws_list[-1]]
# ws = wb[ws_list[1]]

# #############################################################
# # Get Cell Value
# #############################################################
# cell_a1 = ws['A1']
# ic(cell_a1)
# ic(cell_a1.value)
#
# #############################################################
# # Get Formular Cell Value
# #############################################################
# cell_total_lcd = ws['B2593']
# ic(cell_total_lcd.value)
# cell_total_qr = ws['B2594']
# ic(cell_total_qr.value)
#
# #############################################################
# # Get Datetime Cell Value
# #############################################################
# cell_g6 = ws['G6']
# ic(cell_g6.value)
# cell_g6_value = cell_g6.value.strftime('%Y-%m-%d %H:%M:%S')
# ic(cell_g6_value)
#
# #############################################################
# # Get Time Cell Value
# #############################################################
# cell_time = ws['B2596']
# ic(cell_time.value)
# cell_text_time = ws['B2597']
# ic(cell_text_time.value)
#
# #############################################################
# # Get Percent Cell Value
# #############################################################
# cell_ratio = ws['B2599']
# ic(cell_ratio.value)
# cell_text_ratio = ws['B2600']
# ic(cell_text_ratio.value)
#
# #############################################################
# # Get Number Cell Value
# #############################################################
# cell_number = ws['B2602']
# ic(cell_number.value)
# cell_text_number = ws['B2603']
# ic(cell_text_number.value)
# cell_float = ws['B2604']
# ic(cell_float.value)
# cell_with_text = ws['B2605']
# ic(cell_with_text.value)
#
# ############################################################
# Get Cell Color
# ############################################################
# fill_a6 = ws['A6'].fill
# ic(fill_a6.fgColor.index)
# ic(fill_a6.fgColor.type)
# if fill_a6.fgColor.type == 'rgb':
#     argb = fill_a6.fgColor.rgb
#     ic(argb)
#     argb = tuple(int(argb[i:i+2], 16) for i in range(0, len(argb), 2)) # FFFFFFFF
#     ic(argb)
#
# fill_b6 = ws['B6'].fill
# ic(fill_b6.fgColor.index)
# ic(fill_b6.fgColor.type)
# if fill_b6.fgColor.type == 'rgb':
#     argb = fill_b6.fgColor.rgb
#     ic(argb)
#     argb = tuple(int(argb[i:i+2], 16) for i in range(0, len(argb), 2)) # FFFFFFFF
#     ic(argb)

# #############################################################
# # Get All 1st Column
# #############################################################
# index = 6
# while True:
#     cell = f'A{index}' # A6, A7, A8, A9
#     cell_value = ws[cell].value
#
#     if cell_value is None:
#         break
#
#     # ic(cell)
#     ic(cell_value)
#     index += 1

# #############################################################
# # Get All 6th Row
# #############################################################
# index = 0
# while True:
#     cell = f'{chr(ord("A") + index)}6' # ord("A") + 1 = 42 + 1 = 43, chr(43) --> B ==> A6, B6, C6, .....
#     cell_value = ws[cell].value
#
#     if cell_value is None:
#         break
#
#     # ic(cell)
#     ic(cell_value)
#     index += 1

# #############################################################
# # Get column data
# #############################################################
# for col in ws.columns:
#     # value_list = []
#     # for elem in col:
#     #     value_list.append(elem.value)
#     col_valuse = [elem.value for elem in col]
#     ic(col_valuse)
#     break

#############################################################
# Get rows data
#############################################################
# for row in ws.rows:
#     row_valuse = [elem.value for elem in row]
#     ic(row_valuse)
#     break

# for index, row in enumerate(ws.rows):
#     if index > 4:
#         row_valuse = [elem.value for elem in row]
#         ic(row_valuse)
#         break

# #############################################################
# # Get data with iter_rows
# #############################################################
# for row in ws.iter_rows(min_row=6):
#     row_valuse = [elem.value for elem in row]
#     ic(row_valuse)
#     break

# #############################################################
# # Get data with sliding
# #############################################################
# for row in ws['A6': 'J20']:
#     row_valuse = [elem.value for elem in row]
#     ic(row_valuse)

#############################################################
# Check Hidden Row
#############################################################
ic(ws.row_dimensions[6].hidden)
ic(ws.row_dimensions[7].hidden)

#############################################################
# Check Hidden Column
#############################################################
ic(ws.column_dimensions['F'].hidden)
ic(ws.column_dimensions['G'].hidden)